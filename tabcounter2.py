import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import os
import csv
import sys
import re
import unicodedata

# Excel библиотеки
import xlrd  # для XLS
from openpyxl import load_workbook  # для XLSX/XLSM

# Drag & Drop
from tkinterdnd2 import TkinterDnD, DND_FILES


# ====================================================================
#                      Подсчёт вкладок Excel
# ====================================================================
def count_sheets_in_file(path):
    ext = os.path.splitext(path)[1].lower()

    try:
        if ext in [".xlsx", ".xlsm"]:
            wb = load_workbook(path, read_only=True)
            return len(wb.sheetnames)

        elif ext == ".xls":
            wb = xlrd.open_workbook(path)
            return len(wb.sheet_names())

        else:
            return "Неподдерживаемый формат"

    except Exception as e:
        return f"Ошибка: {e}"


# ====================================================================
#                   Получение списка вкладок с индексами
# ====================================================================
def get_sheet_names(path):
    """Возвращает список кортежей (индекс, название вкладки)"""
    ext = os.path.splitext(path)[1].lower()
    
    try:
        if ext in [".xlsx", ".xlsm"]:
            wb = load_workbook(path, read_only=True)
            return [(idx, name) for idx, name in enumerate(wb.sheetnames, 1)]

        elif ext == ".xls":
            wb = xlrd.open_workbook(path)
            return [(idx, name) for idx, name in enumerate(wb.sheet_names(), 1)]

        else:
            return []

    except Exception as e:
        return []


# ====================================================================
#                   Анализ столбцов и заголовков
# ====================================================================
def find_header_row(sheet, max_rows=50):
    """
    Ищет строку с заголовками (как минимум 4 заполненных ячейки подряд)
    Возвращает: (номер_строки, список_заголовков) или (None, [])
    """
    for row_idx in range(1, min(max_rows + 1, sheet.max_row + 1)):
        row_cells = []
        consecutive_filled = 0
        start_col = None
        
        # Проходим по всем колонкам в строке
        for col_idx in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=row_idx, column=col_idx)
            value = cell.value
            
            if value is not None and str(value).strip():
                if start_col is None:
                    start_col = col_idx
                consecutive_filled += 1
                row_cells.append((col_idx, str(value).strip()))
            else:
                # Если был блок заполненных ячеек >= 4, это заголовок
                if consecutive_filled >= 4:
                    return (row_idx, row_cells)
                # Сброс счётчика
                consecutive_filled = 0
                row_cells = []
                start_col = None
        
        # Проверка в конце строки
        if consecutive_filled >= 4:
            return (row_idx, row_cells)
    
    return (None, [])


def find_header_row_xls(sheet, max_rows=50):
    """
    Ищет строку с заголовками для XLS файлов
    Возвращает: (номер_строки, список_заголовков) или (None, [])
    """
    for row_idx in range(min(max_rows, sheet.nrows)):
        row_cells = []
        consecutive_filled = 0
        start_col = None
        
        for col_idx in range(sheet.ncols):
            cell = sheet.cell(row_idx, col_idx)
            value = cell.value
            
            if value is not None and str(value).strip():
                if start_col is None:
                    start_col = col_idx
                consecutive_filled += 1
                row_cells.append((col_idx + 1, str(value).strip()))
            else:
                if consecutive_filled >= 4:
                    return (row_idx + 1, row_cells)
                consecutive_filled = 0
                row_cells = []
                start_col = None
        
        if consecutive_filled >= 4:
            return (row_idx + 1, row_cells)
    
    return (None, [])


def analyze_file_structure(path):
    """
    Анализирует структуру файла: для каждой вкладки находит заголовки
    Возвращает: список [(название_вкладки, количество_столбцов, список_заголовков, номер_строки), ...]
    """
    ext = os.path.splitext(path)[1].lower()
    results = []
    
    try:
        if ext in [".xlsx", ".xlsm"]:
            wb = load_workbook(path, read_only=True, data_only=True)
            
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                header_row, headers = find_header_row(sheet)
                
                if header_row:
                    results.append((sheet_name, len(headers), headers, header_row))
                else:
                    results.append((sheet_name, 0, [], None))
        
        elif ext == ".xls":
            wb = xlrd.open_workbook(path)
            
            for sheet in wb.sheets():
                header_row, headers = find_header_row_xls(sheet)
                
                if header_row:
                    results.append((sheet.name, len(headers), headers, header_row))
                else:
                    results.append((sheet.name, 0, [], None))
        
        return results
    
    except Exception as e:
        print(f"Ошибка анализа файла: {e}")
        return []


def get_column_letter(col_num):
    """Конвертирует номер колонки в буквенное обозначение Excel (1 -> A, 27 -> AA)"""
    result = ""
    while col_num > 0:
        col_num -= 1
        result = chr(65 + (col_num % 26)) + result
        col_num //= 26
    return result


# ====================================================================
#                   Сравнение маппинга столбцов
# ====================================================================
def get_column_signature(headers):
    """
    Создаёт сигнатуру столбцов для сравнения (только названия, без индексов)
    """
    if not headers:
        return None
    # Берём только названия столбцов (игнорируем их позиции)
    return tuple(name.lower().strip() for idx, name in headers)


def group_sheets_by_mapping(structure):
    """
    Группирует вкладки по одинаковому маппингу столбцов
    Возвращает: словарь {signature: [список_индексов_вкладок]}
    """
    mapping_groups = {}
    
    for idx, (sheet_name, col_count, headers, header_row) in enumerate(structure):
        signature = get_column_signature(headers)
        
        if signature is None:
            continue
        
        if signature not in mapping_groups:
            mapping_groups[signature] = []
        
        mapping_groups[signature].append(idx)
    
    return mapping_groups


def get_group_colors():
    """
    Возвращает список цветов для групп вкладок
    """
    return [
        "#C8E6C9",  # Светло-зелёный
        "#E1BEE7",  # Лиловый
        "#FFCCBC",  # Светло-оранжевый
        "#B3E5FC",  # Светло-голубой
        "#FFF9C4",  # Светло-жёлтый
        "#F8BBD0",  # Светло-розовый
        "#D1C4E9",  # Светло-фиолетовый
        "#DCEDC8",  # Светло-салатовый
        "#FFECB3",  # Светло-янтарный
        "#CFD8DC",  # Светло-серый
    ]


# ====================================================================
#                        Управление списком файлов
# ====================================================================
files = []


def add_files():
    file_paths = filedialog.askopenfilenames(
        title="Выберите Excel файлы",
        filetypes=[("Excel files", "*.xlsx *.xlsm *.xls")]
    )
    for path in file_paths:
        if path not in files:
            files.append(path)
            file_list.insert("", tk.END, values=(len(files), path, ""))


def clear_list():
    files.clear()
    for row in file_list.get_children():
        file_list.delete(row)


# ====================================================================
#                     Подсчёт результатов для всех файлов
# ====================================================================
def count_all():
    if not files:
        messagebox.showwarning("Ошибка", "Добавьте хотя бы один файл.")
        return

    results = []
    items = file_list.get_children()
    
    for idx, path in enumerate(files, 1):
        count = count_sheets_in_file(path)
        results.append((idx, os.path.basename(path), count))
        
        if idx - 1 < len(items):
            item_id = items[idx - 1]
            file_list.item(item_id, values=(idx, path, count))

    show_results(results)


# ====================================================================
#                      Показать вкладки выбранного файла
# ====================================================================
def show_sheets():
    selected = file_list.selection()
    if not selected:
        messagebox.showwarning("Ошибка", "Выберите файл из списка.")
        return
    
    item_values = file_list.item(selected[0], 'values')
    file_path = item_values[1]
    file_name = os.path.basename(file_path)
    
    sheets = get_sheet_names(file_path)
    
    if not sheets:
        messagebox.showerror("Ошибка", "Не удалось прочитать вкладки из файла.")
        return
    
    win = tk.Toplevel(root)
    win.title(f"Вкладки файла: {file_name}")
    win.geometry("600x450")
    
    tk.Label(win, text=f"Файл: {file_name}", font=("Arial", 10, "bold")).pack(pady=10)
    tk.Label(win, text=f"Всего вкладок: {len(sheets)}").pack()
    
    table = ttk.Treeview(win, columns=("index", "name"), show="headings", height=15)
    table.heading("index", text="Индекс")
    table.heading("name", text="Название вкладки")
    table.column("index", width=80, anchor="center")
    table.column("name", width=500)
    table.pack(fill="both", expand=True, padx=10, pady=10)
    
    for idx, name in sheets:
        table.insert("", tk.END, values=(idx, name))
    
    def copy_selected():
        selected_item = table.selection()
        if selected_item:
            values = table.item(selected_item[0], 'values')
            sheet_name = values[1]
            win.clipboard_clear()
            win.clipboard_append(sheet_name)
            messagebox.showinfo("Скопировано", f"Название вкладки '{sheet_name}' скопировано в буфер обмена.")
        else:
            messagebox.showwarning("Ошибка", "Выберите вкладку для копирования.")
    
    def copy_all():
        all_names = "\n".join([name for idx, name in sheets])
        win.clipboard_clear()
        win.clipboard_append(all_names)
        messagebox.showinfo("Скопировано", f"Все {len(sheets)} названий вкладок скопированы в буфер обмена.")
    
    btn_frame = tk.Frame(win)
    btn_frame.pack(pady=10)
    
    tk.Button(btn_frame, text="Копировать выбранную", width=20,
              command=copy_selected).grid(row=0, column=0, padx=5)
    
    tk.Button(btn_frame, text="Копировать все названия", width=20,
              command=copy_all).grid(row=0, column=1, padx=5)
    
    tk.Button(btn_frame, text="Сохранить список в CSV", width=20,
              command=lambda: save_sheets_to_csv(file_name, sheets)).grid(row=0, column=2, padx=5)


# ====================================================================
#              Показать все столбцы файла
# ====================================================================
def show_columns():
    selected = file_list.selection()
    if not selected:
        messagebox.showwarning("Ошибка", "Выберите файл из списка.")
        return
    
    item_values = file_list.item(selected[0], 'values')
    file_path = item_values[1]
    file_name = os.path.basename(file_path)
    
    messagebox.showinfo("Анализ", "Анализирую структуру файла...\nЭто может занять несколько секунд.")
    structure = analyze_file_structure(file_path)
    
    if not structure:
        messagebox.showerror("Ошибка", "Не удалось проанализировать структуру файла.")
        return
    
    win = tk.Toplevel(root)
    win.title(f"Структура столбцов: {file_name}")
    win.geometry("700x500")
    
    tk.Label(win, text=f"Файл: {file_name}", font=("Arial", 10, "bold")).pack(pady=10)
    tk.Label(win, text=f"Найдено вкладок: {len(structure)}").pack()
    
    table = ttk.Treeview(win, columns=("sheet", "columns", "header_row"), show="headings", height=12)
    table.heading("sheet", text="Название вкладки")
    table.heading("columns", text="Столбцов")
    table.heading("header_row", text="Строка заголовка")
    table.column("sheet", width=300)
    table.column("columns", width=100, anchor="center")
    table.column("header_row", width=120, anchor="center")
    table.pack(fill="both", expand=True, padx=10, pady=10)
    
    for sheet_name, col_count, headers, header_row in structure:
        row_text = f"Строка {header_row}" if header_row else "Не найдено"
        table.insert("", tk.END, values=(sheet_name, col_count, row_text))
    
    def show_sheet_details():
        selected_item = table.selection()
        if not selected_item:
            messagebox.showwarning("Ошибка", "Выберите вкладку для просмотра деталей.")
            return
        
        item_index = table.index(selected_item[0])
        sheet_name, col_count, headers, header_row = structure[item_index]
        
        if not headers:
            messagebox.showinfo("Информация", f"В вкладке '{sheet_name}' не найдено заголовков\n(нет 4+ заполненных ячеек подряд)")
            return
        
        detail_win = tk.Toplevel(win)
        detail_win.title(f"Столбцы вкладки: {sheet_name}")
        detail_win.geometry("650x450")
        
        tk.Label(detail_win, text=f"Вкладка: {sheet_name}", font=("Arial", 10, "bold")).pack(pady=10)
        tk.Label(detail_win, text=f"Строка заголовка: {header_row} | Всего столбцов: {col_count}").pack()
        
        cols_table = ttk.Treeview(detail_win, columns=("col_num", "col_name"), show="headings", height=15)
        cols_table.heading("col_num", text="Колонка Excel")
        cols_table.heading("col_name", text="Название столбца")
        cols_table.column("col_num", width=120, anchor="center")
        cols_table.column("col_name", width=500)
        cols_table.pack(fill="both", expand=True, padx=10, pady=10)
        
        for col_idx, col_name in headers:
            col_letter = get_column_letter(col_idx)
            cols_table.insert("", tk.END, values=(col_letter, col_name))
        
        btn_frame = tk.Frame(detail_win)
        btn_frame.pack(pady=10)
        
        def copy_columns():
            all_cols = "\n".join([f"{get_column_letter(idx)}: {name}" for idx, name in headers])
            detail_win.clipboard_clear()
            detail_win.clipboard_append(all_cols)
            messagebox.showinfo("Скопировано", f"Все {len(headers)} столбцов скопированы в буфер обмена.")
        
        def save_columns_csv():
            path = filedialog.asksaveasfilename(
                defaultextension=".csv",
                filetypes=[("CSV", "*.csv")],
                initialfile=f"{sheet_name}_columns.csv"
            )
            if not path:
                return
            
            try:
                with open(path, "w", newline="", encoding="utf-8-sig") as f:
                    w = csv.writer(f)
                    w.writerow(["Колонка Excel", "Название столбца"])
                    for col_idx, col_name in headers:
                        w.writerow([get_column_letter(col_idx), col_name])
                messagebox.showinfo("Готово", "Список столбцов сохранён в CSV.")
            except Exception as e:
                messagebox.showerror("Ошибка", str(e))
        
        tk.Button(btn_frame, text="Копировать все столбцы", width=25,
                  command=copy_columns).grid(row=0, column=0, padx=5)
        
        tk.Button(btn_frame, text="Сохранить в CSV", width=25,
                  command=save_columns_csv).grid(row=0, column=1, padx=5)
    
    btn_frame = tk.Frame(win)
    btn_frame.pack(pady=10)
    
    tk.Button(btn_frame, text="Показать столбцы выбранной вкладки", width=35,
              command=show_sheet_details, bg="#2196F3", fg="white", font=("Arial", 9, "bold")).pack()


# ====================================================================
#         НОВАЯ ФУНКЦИЯ: Сравнить маппинг столбцов вкладок
# ====================================================================
def compare_sheet_mappings():
    selected = file_list.selection()
    if not selected:
        messagebox.showwarning("Ошибка", "Выберите файл из списка.")
        return
    
    item_values = file_list.item(selected[0], 'values')
    file_path = item_values[1]
    file_name = os.path.basename(file_path)
    
    messagebox.showinfo("Анализ", "Анализирую и сравниваю маппинг столбцов...\nЭто может занять несколько секунд.")
    structure = analyze_file_structure(file_path)
    
    if not structure:
        messagebox.showerror("Ошибка", "Не удалось проанализировать структуру файла.")
        return
    
    # Группируем вкладки по маппингу
    mapping_groups = group_sheets_by_mapping(structure)
    
    # Фильтруем: оставляем только группы с 2+ вкладками
    filtered_groups = {sig: indices for sig, indices in mapping_groups.items() if len(indices) >= 2}
    
    # Получаем список вкладок без группы (уникальные)
    all_grouped_indices = set()
    for indices in filtered_groups.values():
        all_grouped_indices.update(indices)
    
    unique_indices = [i for i in range(len(structure)) if i not in all_grouped_indices and structure[i][2]]
    
    # Создаём окно результатов
    win = tk.Toplevel(root)
    win.title(f"Сравнение маппинга столбцов: {file_name}")
    win.geometry("800x600")
    
    tk.Label(win, text=f"Файл: {file_name}", font=("Arial", 10, "bold")).pack(pady=10)
    tk.Label(win, text=f"Найдено групп с одинаковым маппингом: {len(filtered_groups)} | Уникальных вкладок: {len(unique_indices)}").pack()
    
    # Создаём стиль для цветных строк
    style = ttk.Style()
    colors = get_group_colors()
    
    # Таблица с вкладками
    table = ttk.Treeview(win, columns=("sheet", "columns", "group"), show="headings", height=18)
    table.heading("sheet", text="Название вкладки")
    table.heading("columns", text="Столбцов")
    table.heading("group", text="Группа")
    table.column("sheet", width=400)
    table.column("columns", width=100, anchor="center")
    table.column("group", width=250, anchor="center")
    table.pack(fill="both", expand=True, padx=10, pady=10)
    
    # Добавляем группированные вкладки с цветами
    group_num = 1
    for idx, (signature, sheet_indices) in enumerate(filtered_groups.items()):
        color = colors[idx % len(colors)]
        
        # Получаем названия столбцов для описания группы
        first_sheet_idx = sheet_indices[0]
        headers = structure[first_sheet_idx][2]
        column_names = ", ".join([name[:20] + "..." if len(name) > 20 else name for _, name in headers[:3]])
        if len(headers) > 3:
            column_names += f" (+{len(headers) - 3} ещё)"
        
        group_label = f"Группа {group_num} ({len(sheet_indices)} вкладок)"
        
        for sheet_idx in sheet_indices:
            sheet_name, col_count, headers, header_row = structure[sheet_idx]
            item_id = table.insert("", tk.END, values=(sheet_name, col_count, group_label), tags=(f"group_{idx}",))
            table.tag_configure(f"group_{idx}", background=color)
        
        group_num += 1
    
    # Добавляем уникальные вкладки без цвета
    for sheet_idx in unique_indices:
        sheet_name, col_count, headers, header_row = structure[sheet_idx]
        table.insert("", tk.END, values=(sheet_name, col_count, "Уникальная"))
    
    # Функция показа деталей группы
    def show_group_details():
        selected_item = table.selection()
        if not selected_item:
            messagebox.showwarning("Ошибка", "Выберите вкладку для просмотра деталей группы.")
            return
        
        values = table.item(selected_item[0], 'values')
        sheet_name = values[0]
        group_label = values[2]
        
        if group_label == "Уникальная":
            messagebox.showinfo("Информация", f"Вкладка '{sheet_name}' имеет уникальный маппинг столбцов.")
            return
        
        # Находим все вкладки в этой группе
        sheet_idx = None
        for idx, (name, _, _, _) in enumerate(structure):
            if name == sheet_name:
                sheet_idx = idx
                break
        
        if sheet_idx is None:
            return
        
        signature = get_column_signature(structure[sheet_idx][2])
        group_sheets = [structure[i] for i in mapping_groups[signature]]
        
        # Показываем окно с деталями группы
        detail_win = tk.Toplevel(win)
        detail_win.title(f"Детали группы: {group_label}")
        detail_win.geometry("700x500")
        
        tk.Label(detail_win, text=f"Группа: {group_label}", font=("Arial", 10, "bold")).pack(pady=10)
        tk.Label(detail_win, text=f"Вкладки в группе: {len(group_sheets)}").pack()
        
        # Список вкладок в группе
        tk.Label(detail_win, text="Вкладки с одинаковым маппингом:", font=("Arial", 9, "bold")).pack(pady=(10, 5))
        sheets_text = tk.Text(detail_win, height=5, width=80)
        sheets_text.pack(padx=10)
        sheets_text.insert("1.0", "\n".join([f"• {s[0]}" for s in group_sheets]))
        sheets_text.config(state="disabled")
        
        # Список столбцов
        tk.Label(detail_win, text="Общий маппинг столбцов:", font=("Arial", 9, "bold")).pack(pady=(10, 5))
        
        cols_table = ttk.Treeview(detail_win, columns=("col_num", "col_name"), show="headings", height=12)
        cols_table.heading("col_num", text="№")
        cols_table.heading("col_name", text="Название столбца")
        cols_table.column("col_num", width=60, anchor="center")
        cols_table.column("col_name", width=600)
        cols_table.pack(fill="both", expand=True, padx=10, pady=10)
        
        headers = group_sheets[0][2]
        for order_num, (col_idx, col_name) in enumerate(headers, 1):
            cols_table.insert("", tk.END, values=(order_num, col_name))
        
        # Кнопка экспорта
        def export_group():
            path = filedialog.asksaveasfilename(
                defaultextension=".csv",
                filetypes=[("CSV", "*.csv")],
                initialfile=f"group_{group_label.replace(' ', '_')}.csv"
            )
            if not path:
                return
            
            try:
                with open(path, "w", newline="", encoding="utf-8-sig") as f:
                    w = csv.writer(f)
                    w.writerow(["Вкладки в группе"])
                    for s in group_sheets:
                        w.writerow([s[0]])
                    w.writerow([])
                    w.writerow(["№", "Название столбца"])
                    for order_num, (col_idx, col_name) in enumerate(headers, 1):
                        w.writerow([order_num, col_name])
                messagebox.showinfo("Готово", "Детали группы сохранены в CSV.")
            except Exception as e:
                messagebox.showerror("Ошибка", str(e))
        
        tk.Button(detail_win, text="Экспортировать группу в CSV", width=30,
                  command=export_group, bg="#4CAF50", fg="white").pack(pady=10)
    
    # Кнопки управления
    btn_frame = tk.Frame(win)
    btn_frame.pack(pady=10)
    
    tk.Button(btn_frame, text="Показать детали группы", width=30,
              command=show_group_details, bg="#FF9800", fg="white", font=("Arial", 9, "bold")).grid(row=0, column=0, padx=5)
    
    def export_all_mappings():
        path = filedialog.asksaveasfilename(
            defaultextension=".csv",
            filetypes=[("CSV", "*.csv")],
            initialfile=f"{os.path.splitext(file_name)[0]}_mappings.csv"
        )
        if not path:
            return
        
        try:
            with open(path, "w", newline="", encoding="utf-8-sig") as f:
                w = csv.writer(f)
                w.writerow(["Название вкладки", "Количество столбцов", "Группа"])
                
                group_num = 1
                for signature, sheet_indices in filtered_groups.items():
                    for sheet_idx in sheet_indices:
                        sheet_name, col_count, _, _ = structure[sheet_idx]
                        w.writerow([sheet_name, col_count, f"Группа {group_num}"])
                    group_num += 1
                
                for sheet_idx in unique_indices:
                    sheet_name, col_count, _, _ = structure[sheet_idx]
                    w.writerow([sheet_name, col_count, "Уникальная"])
            
            messagebox.showinfo("Готово", "Сравнение маппинга сохранено в CSV.")
        except Exception as e:
            messagebox.showerror("Ошибка", str(e))
    
    tk.Button(btn_frame, text="Экспортировать всё в CSV", width=30,
              command=export_all_mappings, bg="#2196F3", fg="white", font=("Arial", 9, "bold")).grid(row=0, column=1, padx=5)


# ====================================================================
#                    Сохранение списка вкладок в CSV
# ====================================================================
def save_sheets_to_csv(file_name, sheets):
    path = filedialog.asksaveasfilename(
        defaultextension=".csv",
        filetypes=[("CSV", "*.csv")],
        initialfile=f"{os.path.splitext(file_name)[0]}_sheets.csv"
    )
    if not path:
        return
    
    try:
        with open(path, "w", newline="", encoding="utf-8-sig") as f:
            w = csv.writer(f)
            w.writerow(["Индекс", "Название вкладки"])
            w.writerows(sheets)
        messagebox.showinfo("Готово", "Список вкладок сохранён в CSV.")
    except Exception as e:
        messagebox.showerror("Ошибка", e)


# ====================================================================
#                             Окно результата
# ====================================================================
def show_results(results):
    win = tk.Toplevel(root)
    win.title("Результаты подсчёта")
    win.geometry("500x350")

    table = ttk.Treeview(win, columns=("num", "file", "count"), show="headings", height=12)
    table.heading("num", text="№")
    table.heading("file", text="Файл")
    table.heading("count", text="Вкладок")
    table.column("num", width=40, anchor="center")
    table.column("file", width=300)
    table.column("count", width=80, anchor="center")
    table.pack(fill="both", expand=True, padx=10, pady=10)

    for r in results:
        table.insert("", tk.END, values=r)

    frame = tk.Frame(win)
    frame.pack(pady=10)

    tk.Button(frame, text="Сохранить в CSV", width=18,
              command=lambda: save_to_csv(results)).grid(row=0, column=0, padx=5)

    tk.Button(frame, text="Сохранить в XLSX", width=18,
              command=lambda: save_to_xlsx(results)).grid(row=0, column=1, padx=5)


# ====================================================================
#                    Сохранение CSV и XLSX
# ====================================================================
def save_to_csv(results):
    path = filedialog.asksaveasfilename(defaultextension=".csv",
                                        filetypes=[("CSV", "*.csv")])
    if not path:
        return

    try:
        with open(path, "w", newline="", encoding="utf-8-sig") as f:
            w = csv.writer(f)
            w.writerow(["№", "Файл", "Количество вкладок"])
            w.writerows(results)
        messagebox.showinfo("Готово", "CSV-файл сохранён.")
    except Exception as e:
        messagebox.showerror("Ошибка", e)


def save_to_xlsx(results):
    path = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                        filetypes=[("Excel", "*.xlsx")])
    if not path:
        return

    try:
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.append(["№", "Файл", "Количество вкладок"])
        for r in results:
            ws.append(r)
        wb.save(path)
        messagebox.showinfo("Готово", "Excel-файл сохранён.")
    except Exception as e:
        messagebox.showerror("Ошибка", e)


# ====================================================================
#       Drag & Drop с ПОЛНОЙ поддержкой всех языков мира
# ====================================================================
def drop(event):
    paths = []
    
    try:
        raw_paths = root.tk.splitlist(event.data)
        
        for path in raw_paths:
            path = path.strip()
            
            for wrapper in ['{}', '""', "''"]:
                if path.startswith(wrapper[0]) and path.endswith(wrapper[1]):
                    path = path[1:-1]
            
            path = path.strip()
            
            if os.path.isfile(path):
                paths.append(path)
                continue
            
            if sys.platform == "win32":
                try:
                    path_normalized = unicodedata.normalize('NFC', path)
                    if os.path.isfile(path_normalized):
                        paths.append(path_normalized)
                        continue
                    
                    path_normalized = unicodedata.normalize('NFD', path)
                    if os.path.isfile(path_normalized):
                        paths.append(path_normalized)
                        continue
                except:
                    pass
                
    except Exception as e:
        raw = event.data.strip()
        
        patterns = [
            r'\{([^}]+)\}',
            r'"([^"]+)"',
            r"'([^']+)'",
        ]
        
        for pattern in patterns:
            matches = re.findall(pattern, raw)
            for p in matches:
                p = p.strip()
                if os.path.isfile(p):
                    paths.append(p)
        
        if not paths:
            for p in raw.split():
                p = p.strip().strip('{}').strip('"').strip("'")
                if os.path.isfile(p):
                    paths.append(p)
    
    added_count = 0
    skipped_count = 0
    
    for path in paths:
        ext = os.path.splitext(path)[1].lower()
        if ext in [".xls", ".xlsx", ".xlsm"]:
            if path not in files:
                files.append(path)
                file_list.insert("", tk.END, values=(len(files), path, ""))
                added_count += 1
        else:
            skipped_count += 1
    
    if added_count > 0:
        msg = f"✅ Добавлено файлов: {added_count}"
        if skipped_count > 0:
            msg += f"\n⚠️ Пропущено (не Excel): {skipped_count}"
        messagebox.showinfo("Результат", msg)
    elif skipped_count > 0:
        messagebox.showwarning("Внимание", f"Пропущено файлов (не Excel): {skipped_count}")


# ====================================================================
#                              GUI
# ====================================================================
root = TkinterDnD.Tk()
root.title("Excel Sheet Counter PRO")
root.geometry("750x600")
root.resizable(False, False)

if sys.platform == "win32":
    try:
        sys.stdout.reconfigure(encoding='utf-8')
    except:
        pass

main = tk.Frame(root, padx=10, pady=10)
main.pack(fill="both", expand=True)

tk.Label(main, text="Перетащите Excel-файлы сюда или нажмите 'Добавить файлы'").pack()

file_list = ttk.Treeview(main, columns=("num", "path", "count"), show="headings", height=10)
file_list.heading("num", text="№")
file_list.heading("path", text="Путь к файлу")
file_list.heading("count", text="Вкладок")
file_list.column("num", width=40, anchor="center")
file_list.column("path", width=580)
file_list.column("count", width=80, anchor="center")
file_list.pack(fill="both", expand=True, pady=10)

file_list.drop_target_register(DND_FILES)
file_list.dnd_bind("<<Drop>>", drop)

btns = tk.Frame(main)
btns.pack()

tk.Button(btns, text="Добавить файлы", width=18, command=add_files).grid(row=0, column=0, padx=5)
tk.Button(btns, text="Очистить список", width=18, command=clear_list).grid(row=0, column=1, padx=5)
tk.Button(btns, text="Подсчитать вкладки", width=18, command=count_all).grid(row=0, column=2, padx=5)

btns2 = tk.Frame(main)
btns2.pack(pady=5)

tk.Button(btns2, text="Показать вкладки выбранного файла", width=40, 
          command=show_sheets, bg="#92D794", fg="white", font=("Arial", 9, "bold")).grid(row=0, column=0, padx=5, pady=2)

tk.Button(btns2, text="Показать все столбцы файла", width=40, 
          command=show_columns, bg="#80CBC4", fg="white", font=("Arial", 9, "bold")).grid(row=1, column=0, padx=5, pady=2)

tk.Button(btns2, text="Сравнить маппинг столбцов вкладок", width=40, 
          command=compare_sheet_mappings, bg="#C290CA", fg="white", font=("Arial", 9, "bold")).grid(row=2, column=0, padx=5, pady=2)

root.mainloop()
